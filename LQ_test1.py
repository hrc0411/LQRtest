#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import io
import base64
import pandas as pd
import dash
from functions import *
#from dash import dash_table
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, State
import dash_bootstrap_components as dbc
import dash_ag_grid as dag
from openpyxl import load_workbook


app = dash.Dash(__name__,external_stylesheets=[dbc.themes.BOOTSTRAP])

app.layout = html.Div([
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and drop or ',
            html.A('select excel file')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px', 
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px',
            'fontWeight': 'bold',
            'color':'blue'
        },
        multiple=False
    ),
    dcc.Dropdown(
        id='select-sheet',
        options=[],
        placeholder='Select the sheet of interest'
    ),
    html.Div([
            html.B('Select product type：'),
            dcc.RadioItems(
                id='select-type',
                options=[
                    {'label': 'IC', 'value': 'IC'},
                    {'label': 'Discrete', 'value': 'Discrete'}
                ],
                value='IC',
                labelStyle={'display': 'inline-block'}
            )
        ],
        style={
            'display': 'flex',
            'flexDirection': 'row',
        },
    ),
    html.Hr(), 
    dbc.Button('Select and Process', id='save-button', n_clicks=0,color="primary",className="me-1"),
    html.Div(id='process-output')
])

    
def save_excel_file(contents, filename):
    _, file_extension = os.path.splitext(filename)
    if file_extension in ['.xls', '.xlsx']:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        with open(filepath, 'wb') as file:
            file.write(contents)
            #file.write(contents.encode('utf8'))
        return True
    else:
        return False


@app.callback(
    Output('select-sheet', 'options'),
    #Output('output-message', 'children'),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename')
)

def parse_excel(contents, filename):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        # Read Excel file
        workbook = load_workbook(io.BytesIO(decoded), data_only=True)
        # Get the sheet names
        sheet_names = workbook.sheetnames
        
        # Return the sheet names as options for the dropdown
        dropdown_options = [{'label': sheet, 'value': sheet} for sheet in sheet_names]
        return dropdown_options

    except Exception as e:
        print(e)
        return []
def update_dropdown_options(contents, filename):
    if contents is not None:
        dropdown_options = parse_excel(contents, filename)
        return dropdown_options
    else:
        return []


@app.callback(
    #Output('output-message', 'children'),
    Output('process-output', 'children'),
    Input('save-button', 'n_clicks'),
    State('select-type', 'value'),
    State('select-sheet', 'value'),
    State('upload-data', 'contents'),
    State('upload-data', 'filename')
    
)
def save_selected_sheet(n_clicks, selected_type, selected_sheet, contents, filename):
    if n_clicks >0 and selected_type is not None and selected_sheet is not None and contents is not None:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        try:
            # Read Excel file
            workbook = load_workbook(io.BytesIO(decoded), data_only=True)
            print("yes")
            # Read the selected sheet
            df = pd.read_excel(workbook, sheet_name=selected_sheet)
            print("yes")
            # Extract data to generate dff2
            attribute_list= ['Assembly Location',
             'Package name',
             'Package type',
             'Number of die',
             'QRC',
             'AEC quality Grade (in case automotive)',
             'Q006 requirement (in case automotive)',
             'Mission profile (within QRC standards)',
             'Packing',
             'Product Type (IC / Discrete)',
             'Package Outline',
             'Basic Type',
             'Wafer Process Line',
             'Wafer Process Site',
             'Wafer Technology',
             'Wafer material',
             'Pre-assembly method',
             'Pre-assembly site',
             'BOAA (Yes / No)',
             'Chip Size',
             'Wafer Size',
             'FSM',
             'BSM',
             'Passivation',
             'BPO',
             'BPP',
             'Leadframe Material',
             'Heat sink Material (if applicable)',
             'LF Supplier',
             'E-pad / non E-pad',
             'Die paddle',
             'E-pad',
             'Lead post (non-fuse)',
             'Lead post (fuse)',
             'Leadframe top plating thickness',
             'Leadframe bottom plating thickness',
             'Lead pitch',
             'Leadframe roughness (in case supplier change)',
             'Core material',
             'Solder Resist',
             'Plating',
             'Substrate Supplier',
             'Substrate thickness',
             'Substrate plating thickness',
             'Die Attach',
             'Material',
             'Capillary Type',
             'Number of down bonds (chip to die pad)',
             'Number of chip to chip bonds',
             'Number of ground bonds (chip to ground ring)',
             'Number of ground bonds (Lead to ground ring)',
             'Wire diameter',
             'Wire bond method',
             'Clip',
             'Pre-assembly bump site',
             'Bump thickness',
             'Material Name',
             'Supplier',
             'Solder ball size',
             'Encapsulant',
             'Adesion promoter',
             'Pre-Assembly Process',
             'FEOL process',
             'BEOL process']

            category_list =["0) Package (or IFX Product)","1) Chip","2.1) Carrier(leadframe)",
                            "2.2) Carrier(Substrate)","3) Die Attach(solder,epoxy,tape...)",
                            "4.1) Interconnect(wire)","4.1.1) Wire bond method",
                            "4.2) Interconnect(Clip)","4.3) Interconnect(Solder Bump)",
                            "4.4) Interconnect(Solder Ball)","5) Encapsulant (resin,glob top, gel,...)",
                            "6) Adhesion promoter", "7) Pre-Assembly Process","8) FEOL process",
                            "9) BEOL process"
                            ]
            category_index =[11,15,12,6,1,7,1,1,2,3,1,1,1,1,1] #number of time repeat
            categories=[]
            for i in range(15):
                categories.extend([category_list[i]] * category_index[i])
                
                
            # list of cells
            reference_cell_list = [
                df.iat[11,4],df.iat[11,5], #0) Package (or IFX Product)
                df.iat[13,4],df.iat[13,5], #1) Chip
                df.iat[14,4],df.iat[14,5], #2.1) Carrier(leadframe)
                df.iat[15,4],df.iat[15,5], #2.2) Carrier(Substrate)
                df.iat[16,4], #3) Die Attach(solder,epoxy,tape...)
                df.iat[17,4],df.iat[17,5], #4.1) Interconnect(wire)
                df.iat[18,4], #4.1.1) Wire bond method
                df.iat[19,4], #4.2) Interconnect(Clip)
                df.iat[20,4],df.iat[20,5], #4.3) Interconnect(Solder Bump)
                df.iat[21,4],df.iat[21,5], #4.4) Interconnect(Solder Ball)
                df.iat[22,4], #5) Encapsulant (resin,glob top, gel,...)
                df.iat[23,4], #6) Adhesion promoter
                df.iat[24,4], #7) Pre-Assembly Process
                df.iat[25,4], #8) FEOL process
                df.iat[26,4], #9) BEOL process
            ]

            target_cell_list = [
                df.iat[11,13],df.iat[11,14], #0) Package (or IFX Product)
                df.iat[13,13],df.iat[13,14], #1) Chip
                df.iat[14,13],df.iat[14,14], #2.1) Carrier(leadframe)
                df.iat[15,13],df.iat[15,14], #2.2) Carrier(Substrate)
                df.iat[16,13], #3) Die Attach(solder,epoxy,tape...)
                df.iat[17,13],df.iat[17,14], #4.1) Interconnect(wire)
                df.iat[18,13], #4.1.1) Wire bond method
                df.iat[19,13], #4.2) Interconnect(Clip)
                df.iat[20,13],df.iat[20,14], #4.3) Interconnect(Solder Bump)
                df.iat[21,13],df.iat[21,14], #4.4) Interconnect(Solder Ball)
                df.iat[22,13], #5) Encapsulant (resin,glob top, gel,...)
                df.iat[23,13], #6) Adhesion promoter
                df.iat[24,13], #7) Pre-Assembly Process
                df.iat[25,13], #8) FEOL process
                df.iat[26,13], #9) BEOL process
            ]

            # Initialize an empty list to store extracted data for each cell
            headers = []
            reference_values = []
            target_values = []

            # Extract reference data in the list
            for reference_cell_content in reference_cell_list:
                # Split the cell content into lines
                reference_lines = reference_cell_content.strip().split('\n')
                

                # Loop through each line and extract values after the colons
                for line in reference_lines:
                    parts = line.split(':')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        reference_value = parts[1].strip()
                        headers.append(key)
                        reference_values.append(reference_value)
                    else:
                        headers.append("Pending")
                        reference_values.append(line.strip())
                        
            # Extract target data in the list
            for target_cell_content in target_cell_list:
                # Split the cell content into lines
                target_lines = target_cell_content.strip().split('\n')
                
                # Loop through each line and extract values after the colons
                for line in target_lines:
                    parts = line.split(':')
                    if len(parts) == 2:
                        target_value = parts[1].strip()
                        target_values.append(target_value)
                    else:
                        target_values.append(line.strip())
                        
            # Create a DataFrame from the list of extracted data
            data_dict = {'Category': categories, 'Attribute':attribute_list,
                         'Reference': reference_values, 'Target': target_values}
            dff2 = pd.DataFrame(data_dict)
            dff2["Result"] = dff2.apply(generate_result, axis=1)
            dff2[["TC","HTSL","HTRB","HTGB","THB","HAST","H3TRB","PTC","IOL","UHAST"]] = dff2.apply(generate_test, axis=1)
            ###################################


            if selected_type == 'IC':    
                #dff = dff2.copy()
                ########generate summary table###########
                dff3 = dff2.iloc[:,5:]
                summary = pd.DataFrame()
                for col in dff3.columns:
                    value_counts = dff3[col].value_counts().reset_index()
                    value_counts['index'] = value_counts['index'].astype(str)
                    # join all symbol together
                    symbols = ', '.join(value_counts['index'])
                    # if X is included, only show X otherwise everything
                    if "X" in symbols:
                        summary[col] = ["X"]
                    else:
                        summary[col] = symbols

                # delete the empty colomn
                summary = summary.dropna(axis=1, how='all')
                column_defs=[{"headerName": col, "field": col} for col in summary.columns] #for aggrid table
            else:
                dff = pd.read_excel(r"C:\Users\HuangRunchen\Documents\Training\Lean Qual Referencing\data extract1.xlsm", sheet_name="Comparison process(DISCRETE)")
                summary = pd.read_excel(r"C:\Users\HuangRunchen\Documents\Training\Lean Qual Referencing\data extract1.xlsm", sheet_name="summary(IC)")
            
                        
            ########################return dashtable#############################    
            return html.Div([
                dag.AgGrid(
                    id='table1',
                    rowData=dff2[:].to_dict('records'),
                    columnDefs=[
                        {
                            "field": "Category",
                            "rowSpan": {"function": "rowSpanningSimple(params)"},
                            "cellClassRules": {
                                "spanned-row": "params.value==='0) Package (or IFX Product)' || params.value==='1) Chip'|| params.value==='2.1) Carrier(leadframe)'",
                            },  # "0) Package (or IFX Product)","1) Chip","2.1) Carrier(leadframe)",
                                           
                        },
                        {"field": "Attribute"},
                        {"field": "Reference"},
                        {"field": "Target"},
                        {"field": "Result"},
                        {"field": "TC" ,"width":100},
                        {"field": "HTSL","width":100},
                        {"field": "HTRB","width":100},
                        {"field": "HTGB","width":100},
                        {"field": "THB","width":100},
                        {"field": "HAST","width":100},
                        {"field": "H3TRB","width":100},
                        {"field": "PTC","width":100},
                        {"field": "IOL","width":100},
                        {"field": "UHAST","width":100}
                    ],
                    defaultColDef={"resizable": True,"filter": True},
                    columnSize="autoSize",
                    dashGridOptions={"suppressRowTransform": True},
                ),
                html.Hr(),
                html.H2('Recommended Stress Test'),
                html.Div(
                dag.AgGrid(
                        columnDefs=column_defs,                       
                        id='table2',
                        rowData=summary[:].to_dict('records'),
                        defaultColDef={"resizable": True, "minWidth" :50},
                        columnSize="sizeToFit",
                        dashGridOptions={"suppressRowTransform": True},
                    ),
                )
            ])
        except Exception as e:
            print(e)
            return html.Div('error')
    else:
        return html.Div('Please upload an Excel file and select a sheet')



if __name__ == '__main__':
    app.run_server(debug=True)



# In[35]:


#xls = r"C:\Users\HuangRunchen\Documents\Training\Lean Qual Referencing\data extract1.xlsm"
#df = pd.read_excel(xls, sheet_name='Delta Assessment (Trial 1)')
#df


# In[26]:


# In[36]:

#Extract reference data


# In[34]:



# In[35]:




 # In[37]:

#!/usr/bin/env python
# coding: utf-8

# In[5]:


import os
import io
import base64
import pandas as pd
import dash
from dash import dash_table
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, State
import dash_bootstrap_components as dbc
import xlwings as xw
from openpyxl import load_workbook

UPLOAD_FOLDER = r'C:\temp'
SAVE_FOLDER = r'C:\temp'
#if not os.path.exists(UPLOAD_FOLDER):
    #os.makedirs(UPLOAD_FOLDER)
app = dash.Dash(__name__,external_stylesheets=[dbc.themes.BOOTSTRAP])

app.layout = html.Div([
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and drop or ',
            html.A('select excel file')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px', 
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px',
            'fontWeight': 'bold',
            'color':'blue'
        },
        multiple=False
    ),
    dcc.Dropdown(
        id='select-sheet',
        options=[],
        placeholder='Select the sheet of interest'
    ),
    html.Div([
            html.B('Select product type：'),
            dcc.RadioItems(
                id='select-type',
                options=[
                    {'label': 'IC', 'value': 'IC'},
                    {'label': 'Discrete', 'value': 'Discrete'}
                ],
                value='IC',
                labelStyle={'display': 'inline-block'}
            )
        ],
        style={
            'display': 'flex',
            'flexDirection': 'row',
        },
    ),
    html.Hr(), 
    dbc.Button('Select and Process', id='save-button', n_clicks=0,color="primary",className="me-1"),
    #html.Div(id='output-message'),
    #html.Button('Processing', id='process-button', n_clicks=0),
    html.Div(id='process-output')
])

    
def save_excel_file(contents, filename):
    _, file_extension = os.path.splitext(filename)
    if file_extension in ['.xls', '.xlsx']:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        with open(filepath, 'wb') as file:
            file.write(contents)
            #file.write(contents.encode('utf8'))
        return True
    else:
        return False


@app.callback(
    Output('select-sheet', 'options'),
    #Output('output-message', 'children'),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename')
)

def parse_excel(contents, filename):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        # Read Excel file
        workbook = load_workbook(io.BytesIO(decoded), data_only=True)
        # Get the sheet names
        sheet_names = workbook.sheetnames
        
        # Return the sheet names as options for the dropdown
        dropdown_options = [{'label': sheet, 'value': sheet} for sheet in sheet_names]
        return dropdown_options

    except Exception as e:
        print(e)
        return []
def update_dropdown_options(contents, filename):
    if contents is not None:
        dropdown_options = parse_excel(contents, filename)
        return dropdown_options
    else:
        return []


@app.callback(
    #Output('output-message', 'children'),
    Output('process-output', 'children'),
    Input('save-button', 'n_clicks'),
    State('select-type', 'value'),
    State('select-sheet', 'value'),
    State('upload-data', 'contents'),
    State('upload-data', 'filename')
    
)
def save_selected_sheet(n_clicks, selected_type, selected_sheet, contents, filename):
    if n_clicks >0 and selected_type is not None and selected_sheet is not None and contents is not None:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        try:
            # Read Excel file
            workbook = load_workbook(io.BytesIO(decoded), data_only=True)
            # Read the selected sheet
            sheet = workbook[selected_sheet]
            # Convert the sheet to a DataFrame
            df = pd.DataFrame(sheet.values)
            # Extract column headers from the first row
            df.columns = df.iloc[0]
            df = df[1:]
            
            #output_directory = 'C:\temp'
            #file_name = 'output_Delta Assessment.xlsx'
            #output_path = os.path.join(output_directory, file_name)
            df.to_excel(r"C:\temp\output_Delta Assessment.xlsx",index=False)
            ###################################
            # Open the Excel file with xlwings
            xlwings_json_config = {
                "file_path": r"C:\temp\LQreferencing.xlsm",
                "macros": ["Clearcontent.ClearContentExamples", "Module1.test"],
            }
            wb = xw.Book(json=xlwings_json_config)
            wb.save()
            if selected_type == 'IC':    
                dff = pd.read_excel(r"C:\temp\LQreferencing.xlsm", sheet_name="Comparison process(IC)")
                summary = pd.read_excel(r"C:\temp\LQreferencing.xlsm", sheet_name="summary(IC)")
            else:
                dff = pd.read_excel(r"C:\temp\LQreferencing.xlsm", sheet_name="Comparison process(DISCRETE)")
                summary = pd.read_excel(r"C:\temp\LQreferencing.xlsm", sheet_name="summary(DISCRETE)")
            if len(wb.app.books) == 1:
                wb.app.quit()
            else:
                wb.close()
                
            #######################dash table formatting######################    
            style_cell_conditional = []
            style_data_conditional = [
                {
                    'if': {'filter_query': '{Result} = "no"'},
                    'color': 'red'
                }
            ]
            for col in dff.columns:
                if col in ['Result', 'TC', 'HTSL', 'HTRB', 'HTGB', 'THB', 'HAST', 'H3TRB', 'PTC', 'IOL', 'UHAST']:
                    style_cell_conditional.append({'if': {'column_id': col}, 'textAlign': 'center','fontSize':20})
                else:
                    style_cell_conditional.append({'if': {'column_id': col}, 'textAlign': 'left','fontSize':20})
            #######################dash table formatting######################         
            ########################return dashtable#############################    
            return html.Div([
                dash_table.DataTable(
                    id='table1',
                    columns=[{'name': col, 'id': col} for col in dff.columns],
                    data=dff[:-1].to_dict('records'),
                    style_cell_conditional=style_cell_conditional,
                    style_data_conditional=style_data_conditional,
                    style_header={'textAlign': 'center','backgroundColor': 'yellow','fontWeight': 'bold'},
                    editable=True,
                    filter_action='native',
                    #sort_action='native',
                    #sort_mode='multi',
                    column_selectable='single',
                    #row_selectable='multi',
                    row_deletable=False,
                    selected_columns=[],
                    selected_rows=[],
                    page_action='none',
                ),
                html.Hr(),
                html.H2('Recommended Stress Test'),
                html.Div(
                    dash_table.DataTable(
                        id='table2',
                        #columns=[{'name': col, 'id': col} for col in summary_IC.columns],
                        data=summary.to_dict('records'),
                        style_cell={'textAlign': 'left','fontSize': 30},
                        style_header = {'display': 'none'},
                        editable=True,
                        column_selectable='single',
                        #row_selectable='multi',
                        row_deletable=False,
                        selected_columns=[],
                        selected_rows=[],
                        page_action='none',                       
                    ),
                    style={'overflowX': 'auto'}
                )
            ])
        except Exception as e:
            print(e)
            return html.Div('Excel VBA error')
    else:
        return html.Div('Please upload an Excel file and select a sheet')



if __name__ == '__main__':
    app.run_server(debug=True)



# In[ ]:




